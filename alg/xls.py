import openpyxl

file = "D:\Document\青岛\财务接口需求\data.xlsx"
inwb = openpyxl.load_workbook(file)

sheetnames = inwb.get_sheet_names()

sheet1 = inwb.get_sheet_by_name(sheetnames[0])

rows = sheet1.max_row
columns = sheet1.max_column

pluMap = {}

for i in range(rows):

    plu = sheet1.cell(int(i + 1), 1).value
    pluname = sheet1.cell(int(i + 1), 2).value
    pluTaxClass = sheet1.cell(int(i + 1), 3).value

    r = pluMap.get(pluTaxClass)
    if r is None:
        pluMap.update({pluTaxClass:[[plu, pluname]]})
    else:
        pluMap.get(pluTaxClass).append([plu, pluname])

TEMP1 = """INSERT INTO XF_ITEMSINVOICEH VALUES ('{TAXCLASSCODE:}','');"""

TEMP = """INSERT INTO XF_ITEMSINVOICED (XF_ITEMORGID,XF_PLU,XF_TAXCLASSCODE,XF_PLUDESC,XF_ITEMORGDESC)
SELECT ML.XF_ITEMORGID, ML.XF_PLU, '{TAXCLASSCODE:}' XF_TAXCLASSCODE , TE.XF_NAME, MS.XF_DESCI
			  FROM XF_ITEMMAS MS, XF_ITEMLOOKUP ML,XF_MDTENANT TE
			 WHERE MS.XF_ITEMORGID = ML.XF_ITEMORGID
			   AND MS.XF_STYLE = ML.XF_PLU AND ML.XF_ITEMORGID = TE.XF_ORGID
         AND ML.XF_PLU = '{PLU:}'
			   AND NOT EXISTS (SELECT 1 FROM XF_ITEMSINVOICED T WHERE T.XF_PLU = ML.XF_PLU AND T.XF_ITEMORGID = ML.XF_ITEMORGID );
		"""

keyCahe = set()

for k,v in pluMap.items():
    # 生成sql
    for i in v:
        if k is not None:
            if k not in keyCahe:
                print(TEMP1.format(TAXCLASSCODE=k))
                keyCahe.add(k)

            print(TEMP.format(PLU=i[0], TAXCLASSCODE=k))

